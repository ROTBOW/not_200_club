import os
from datetime import date, datetime
from statistics import mean, median, mode

import openpyxl

# Constants
DIR = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
TARGET = os.path.join(DIR, 'target')
RES = os.path.join(DIR, 'res')

class DTX:
    
    def __init__(self) -> None:
        self.workbook = openpyxl.Workbook()
        self.workbook.remove(self.workbook.active)
        self.output_path = os.path.join(RES, f'{"not200club "+str(date.today())}.xlsx')
        self.start_time = datetime.now().strftime("%m/%d/%Y, %H:%M:%S")
        
        
    def save(self) -> None:
        """
        The `save` function saves the workbook to the specified output path.
        """
        self.workbook.save(self.output_path)
       
       
    def __fit_to_data(self, sheet):
        """
        The function `__fit_to_data` adjusts the width of columns in a sheet based on the length of the
        data in each column, with a maximum width limit of 200 characters.
        
        :param sheet: The `sheet` parameter in the `__fit_to_data` method is expected to be a worksheet
        object where the data is stored. This method is designed to adjust the column widths in the
        worksheet based on the length of the data in each column
        """
        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            length = min(length, 200)  # Limit column width to 200
            sheet.column_dimensions[column_cells[0].column_letter].width = length 
      
      
    def __create_coach_sheet(self, coach:str):
        """
        The function creates a new sheet in a workbook with specified bolded column headers and returns the
        sheet.
        
        :param coach: The "coach" parameter is a string that represents the name of the coach for whom
        the sheet is being created
        :type coach: str
        :return: a sheet object.
        """
        sheet = self.workbook.create_sheet(title=coach)
        
        bold_font = openpyxl.styles.Font(bold=True)
        
        for index, header in enumerate(["Seeker Name", "Seeker Status", "Solo Issues", "Capstone Issues", "Group Issues"], start=1):
            sheet.cell(row=1, column=index, value=header).font = bold_font
        
        return sheet
        
        
    def write_coach_sheet(self, coach:str, issues:list, sites_by_coach:dict, overview:dict) -> None:
        """
        The function `write_coach_sheet` writes information about seekers, their issues, and status to a
        coach's sheet in an Excel workbook.
        
        :param coach: The `coach` parameter is a string representing the name of the coach for whom the
        coach sheet is being generated
        :type coach: str
        :param issues: The `issues` parameter in the `write_coach_sheet` method is expected to be a
        dictionary where the keys are seekers and the values are dictionaries containing project-related
        issues. The structure of the `issues` dictionary should be like this:
        :type issues: list
        :param sites_by_coach: The `sites_by_coach` parameter is a dictionary that contains information
        about the status of seekers assigned to a specific coach. The keys in this dictionary are coach
        names, and the values are dictionaries with seeker names as keys and status information as
        values. The status information could include details such as the seeker
        :type sites_by_coach: dict
        :param overview: The `overview` parameter in the `write_coach_sheet` method is a dictionary that
        keeps track of the number of seekers with issues. It is updated within the loop to increment the
        count of seekers with issues. This count is used for generating an overview of the data at the
        end of the method
        :type overview: dict
        """
        row = 2
        sheet = self.__create_coach_sheet(coach)
        
        for seeker, issues in issues.items():
            status = sites_by_coach[coach][seeker]['status']
            overview['seeker_with_issue'] += 1
            
            col = 3
            sheet.cell(row=row, column=1, value=seeker)
            sheet.cell(row=row, column=2, value=status)
            for proj in ['solo', 'capstone', 'group']:
                if issues[proj]:
                    val = '\n'.join([f"{key}: {value}" for key, value in issues[proj].items()])
                else:
                    val = 'No Issues Found'
                    
                sheet.cell(row=row, column=col, value=val)
                col += 1
            row += 1
                
        self.__fit_to_data(sheet)
            
        self.workbook.save(self.output_path)
        
    def fill_overview(self, overview:dict, total_seekers:int, timeout:int) -> None:
        """
        This function fills in an Excel sheet with various statistics related to website loading issues.
        """
        sheet = self.workbook.create_sheet(title='Overview', index=0)
        
        sheet.cell(row=1, column=1, value='TOTAL SEEKERS WITH ISSUES')
        sheet.cell(row=1, column=2, value=f"{overview['seeker_with_issue']}/{total_seekers}")
        sheet.cell(row=1, column=4, value=f'Script ran at {self.start_time} for this sheet')
        
        sheet.cell(row=2, column=1, value='SITES WITH TIMES 10s>')
        sheet.cell(row=2, column=2, value=f'Total: {sum(overview["sites_time"].values())}')
        sheet.cell(row=2, column=3, value=f'Solo: {overview["sites_time"]["solo"]}')
        sheet.cell(row=2, column=4, value=f'Capstone: {overview["sites_time"]["capstone"]}')
        sheet.cell(row=2, column=5, value=f'Group: {overview["sites_time"]["group"]}')
        
        sheet.cell(row=3, column=1, value='LOADING TIME STATS')
        if overview['time_average']:
            sheet.cell(row=3, column=2, value=f'Mean: {round(mean(overview["time_average"]), 2)}s')
            sheet.cell(row=3, column=3, value=f'Mode: {round(mode(overview["time_average"]), 2)}s')
            sheet.cell(row=3, column=4, value=f'Median: {round(median(overview["time_average"]), 2)}s')
        else:
            sheet.cell(row=3, column=2, value='No loading issues found')
            
        
        if timeout:
            sheet.cell(row=4, column=1, value='SITES THAT TIMEOUT')
            sheet.cell(row=4, column=2, value=f"Total: {sum(overview['sites_timeout'].values())}")
            sheet.cell(row=4, column=3, value=f"Solo: {overview['sites_timeout']['solo']}")
            sheet.cell(row=4, column=4, value=f"Capstone: {overview['sites_timeout']['capstone']}")
            sheet.cell(row=4, column=5, value=f"Group: {overview['sites_timeout']['group']}")
        else:
            sheet.cell(row=4, column=1, value='TIMEOUT NOT SET')
            
        for row, group, key in [(5, 'SITES WITH NO URLS', 'sites_no_url'), (6, "SITES WITH BAD URLS", 'sites_bad_url'), (7, 'SITES WITH BAD STATUS', 'sites_status')]:
            sheet.cell(row=row, column=1, value=group)
            sheet.cell(row=row, column=2, value=f"Total: {sum(overview[key].values())}")
            sheet.cell(row=row, column=3, value=f"Solo: {overview[key]['solo']}")
            sheet.cell(row=row, column=4, value=f"Capstone: {overview[key]['capstone']}")
            sheet.cell(row=row, column=5, value=f"Group: {overview[key]['group']}")
            
        self.__fit_to_data(sheet)
        
    def fill_issue_legend(self, timeout:int) -> None:
        """
        This function creates a legend sheet in a workbook with explanations for different types of
        issues that can occur while checking a website.
        """
        sheet = self.workbook.create_sheet(title='Issue Legend', index=0)
        
        sheet.cell(row=1, column=1, value='Issue Type')
        sheet.cell(row=1, column=2, value="Explained")
        sheet.cell(row=1, column=3, value=f'Script ran at {self.start_time} for this sheet')
        
        sheet.cell(row=2, column=1, value='Status')
        sheet.cell(row=2, column=2, value='Will most likely be a 404 or a 503 - both mean the site is down')
        
        sheet.cell(row=3, column=1, value='Bad URL')
        sheet.cell(row=3, column=2, value='The site\'s URL doesn\'t work, The script was unable to even try to check it')
        
        sheet.cell(row=4, column=1, value='Time')
        sheet.cell(row=4, column=2, value='The amount of time in seconds it took to get a response from the site - it needs to have taken longer than 10s to be listed')
        
        sheet.cell(row=5, column=1, value='Timeout')
        sheet.cell(row=5, column=2, value=f'The site took longer than the given timeout({timeout}s) and gave up on the site')
        
        sheet.cell(row=6, column=1, value='No-link')
        sheet.cell(row=6, column=2, value='Means there was no url listed in saleforce for that project')

        self.__fit_to_data(sheet)