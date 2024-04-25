import json
import os
import re
from collections import defaultdict as ddict
from concurrent.futures import ThreadPoolExecutor
from datetime import date, timedelta
from time import time
from warnings import simplefilter

import openpyxl
import requests
from alive_progress import alive_bar

from upload import Uploader
from util.data_to_xlsx import DTX

# Constants
DIR = os.path.dirname(os.path.realpath(__file__))
TARGET = os.path.join(DIR, 'target')
RES = os.path.join(DIR, 'res')

class Not200Club:
    
    @classmethod
    def validate(cls) -> None:
        """
        The function `validate` checks if the necessary folders exist, if the target folder is not
        empty, if the first file in the target folder is not a `.DS_Store` file, and if the file in the
        target folder is an `.xlsx` file.
        """
        
        # Creating output(res) folder if it doesn't exist
        if not os.path.isdir(RES):
            os.mkdir(RES)
            
        # Creating target folder if it doesn't exist
        if not os.path.isdir(TARGET):
            os.mkdir(TARGET)
            
        # Check that the target file has is not empty
        def empty_check() -> None:
            if len(os.listdir(TARGET)) == 0:
                raise Exception('TARGET ERROR - The target folder needs to have one report file in it!')
        empty_check()
        
        # Check if the first file in target is ds_store and remove it if so
        if os.listdir(TARGET)[0] == '.DS_Store':
            os.remove(os.path.join(TARGET, '.DS_Store'))
            # check if empty again if we remove the ds_store
            empty_check()
            
        # Checks that the target file is an xlsx file
        if not os.listdir(TARGET)[0].endswith('.xlsx'):
            raise Exception(f'INVALID FILE TYPE ERROR - The current file in the target folder is a bad type! it\'s not a xlsx file!\nTrying to read: ({os.listdir(TARGET)[0]})')
    
    def __init__(self, timeout = None) -> None:
        """
        The function initializes various data structures and variables for tracking statistics related
        to coaching sites and seekers.
        
        :param timeout: The `timeout` parameter in the `__init__` method is used to specify a timeout
        value for the instance. This value can be used to set a time limit for certain operations or
        processes within the class. If a timeout value is provided when initializing an instance of the
        class, it will be
        """
        self.dtx = DTX()
        self.data = list()
        self.sites_by_coach = ddict(lambda: ddict(dict))
        self.all_coach_issues = dict()
        self.timeout = timeout
        self.total_seekers = 0
        
        
        overview_init = lambda: {'solo': 0, 'capstone': 0, 'group': 0}
        self.overview = {
            'sites_status': overview_init(),
            'sites_time': overview_init(),
            'sites_bad_url': overview_init(),
            'sites_no_url': overview_init(),
            'seeker_with_issue': 0,
            'time_average': list(),
            'sites_timeout': overview_init()
        }
        
    def __validation_check(self) -> bool:
        """
        The function checks if a report file is older than five days and prompts the user to pull a new
        report if necessary.
        :return: a boolean value. If the user's response is 'yes' or 'y', it will return True. If the
        user's response is 'no' or 'n', it will return False.
        """
        file_mod_time = os.path.getmtime(os.path.join(TARGET, f'{os.listdir(TARGET)[0]}'))
        file_age = timedelta(seconds=time()-file_mod_time)
        
        if file_age > timedelta(days=5):
            print('Report File is older than five days, consider pulling a new report\nRun script anyways?')
            while True:
                ans = input('-> ')
                if ans.lower() in ['yes', 'y']:
                    return True
                elif ans.lower() in ['no', 'n']:
                    return False
                print('unknown response, try again\n(ans include y, n)')
                
        return True
        
    def __grab_data_from_file(self) -> None:
        """
        This function reads data from an Excel file in the target folder and populates a dictionary with the data.
        """
        target_file = os.listdir(TARGET)[0]
        
        simplefilter("ignore") # both simplefilter lines supress the style warning openpyxl throws
        data = openpyxl.load_workbook(os.path.join(TARGET, target_file))
        simplefilter("default")
        
        sheet = data.active
        self.total_seekers = sheet.max_row-1

        with alive_bar(sheet.max_row-1, title="Grabing Data...") as bar:
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                curr_row = dict()
                for idx, cell in enumerate(row):
                    
                    data_list = {
                        0: 'seeker',
                        1: 'coach',
                        2: 'status',
                        3: 'solo',
                        4: 'capstone',
                        5: 'group',
                        6: 'email'
                    }
                    
                    val = cell.value
                    if val == ' ' and idx == 1:
                        val = 'Placements' 
                    
                    curr_row[data_list[idx]] = val
                    
                        
                
                for proj in ['status', 'solo', 'capstone', 'group', 'email']:
                    self.sites_by_coach[curr_row['coach']][curr_row['seeker']][proj] = curr_row[proj]
                bar()
    
    def __validate_url(self, url: str) -> str:
        """
        This function validates a given URL and adds "https://" if it is missing.
        
        :param url: The `url` parameter is a string that represents a URL
        :type url: str
        :return: a string that represents a validated URL. If the input URL is empty, it returns an
        empty string. If the input URL does not start with "http://" or "https://", it adds "https://"
        to the beginning of the URL and returns it. Otherwise, it returns the input URL as is.
        """
        if not url:
            return ''
        
        if not re.match(r'^http[s]?:\/\/', url):
            return 'https://' + url
        
        return url
    
    def __get_issues_from_urls(self, seeker: str, urls: dict) -> dict:
        """
        This function takes in a seeker and a dictionary of URLs, and returns a dictionary of issues for
        each URL, along with an overview of the number of issues encountered.
        
        :param seeker: The parameter "seeker" is a string that represents the name of the seeker or the
        person who is running the code
        :type seeker: str
        :param urls: The `urls` parameter is a dictionary containing project names as keys and their
        corresponding URLs as values
        :type urls: dict
        :return: a dictionary containing the issues found in the provided URLs. If no issues are found,
        an empty dictionary is returned. The dictionary is nested within another dictionary with the key
        'seeker'.
        """
        site_issues = ddict(dict)
        
        for proj, url in urls.items():
            if url != '':
                try:
                    res = requests.get(url, timeout=self.timeout)
                    if res.elapsed > timedelta(seconds=10):
                        site_issues[proj]['time'] = res.elapsed
                        self.overview['sites_time'][proj] += 1
                        self.overview['time_average'].append(res.elapsed.seconds)
                                
                    if res.status_code != 200:
                        site_issues[proj]['status'] = res.status_code
                        self.overview['sites_status'][proj] += 1
                        
                except requests.exceptions.Timeout:
                    site_issues[proj]['timeout'] = f'URL timeout at {self.timeout}s'
                    self.overview['sites_timeout'][proj] += 1
                    
                except Exception as e:
                    site_issues[proj]['bad_url'] = str(e)
                    self.overview['sites_bad_url'][proj] += 1
            else:
                site_issues[proj]['no-link'] = True
                self.overview['sites_no_url'][proj] += 1
            
        return {seeker: site_issues} if site_issues else {}
    
    def __threading_get_all_issues(self, coach: str) -> dict:
        """
        This function uses multithreading to gather all issues from URLs associated with a given coach.
        
        :param coach: The coach parameter is a string representing the name of the coach for whom we
        want to get all the issues
        :type coach: str
        :return: a dictionary containing all the issues gathered from the URLs of the seekers assigned
        to the given coach.
        """
        seekers = self.sites_by_coach[coach]
        all_issues = dict()
        
        with ThreadPoolExecutor(max_workers=(os.cpu_count() or 1) * 2) as pool:
            futures = list()
            for seeker in seekers:
                urls = {
                        'solo': self.__validate_url(seekers[seeker]['solo']),
                        'capstone': self.__validate_url(seekers[seeker]['capstone']),
                        'group': self.__validate_url(seekers[seeker]['group'])
                    }
                futures.append(pool.submit(self.__get_issues_from_urls, seeker, urls))
            
            with alive_bar(len(futures), title=f"Gathering threads for {coach}") as bar:
                for f in futures:
                    all_issues.update(f.result())
                    bar()
        
        return all_issues
    
    
    def __test_urls_and_write_to_xlsx(self) -> None:
        """
        The function iterates through a list of coaches, retrieves issues for each coach, and writes the
        information to an Excel file.
        """

        # for coach in ['Josiah Leon']:
        for coach in self.sites_by_coach:
            self.all_coach_issues[coach] = self.__threading_get_all_issues(coach)
            self.dtx.write_coach_sheet(coach, self.all_coach_issues[coach], self.sites_by_coach, self.overview)
        
    def __output_json(self) -> None:
                        
        for coach in self.all_coach_issues:
            for seeker in self.all_coach_issues[coach]:
                self.all_coach_issues[coach][seeker]['email'] = self.sites_by_coach[coach][seeker]['email']
                for proj in self.all_coach_issues[coach][seeker]:
                    if proj == 'email':
                            continue
                    for issue in self.all_coach_issues[coach][seeker][proj]:
                        if isinstance(self.all_coach_issues[coach][seeker][proj][issue], timedelta):
                            self.all_coach_issues[coach][seeker][proj][issue] = str(self.all_coach_issues[coach][seeker][proj][issue].total_seconds())
        
        with open(os.path.join(RES, f'{"not200club "+str(date.today())}.json'), 'w') as file:
            json.dump(self.all_coach_issues, file)
            
    def __upload_json(self) -> None:
        """
        This method uploads the most recent JSON data to firebase.
        """
        up = Uploader()
        up.upload_data()
    
    def main(self) -> None:
        """
        The main function performs various tasks including validation checks, grabbing data from a file,
        filling an issue legend, testing URLs and writing to an Excel file, and filling an overview.
        """
        if self.__validation_check():
            self.__grab_data_from_file()
            self.__test_urls_and_write_to_xlsx()
            self.dtx.fill_issue_legend(self.timeout)
            self.dtx.fill_overview(self.overview, self.total_seekers, self.timeout)
            self.dtx.save()
            self.__output_json()
            self.__upload_json()
 

if __name__ == '__main__':
    Not200Club.validate()

    start = time()
    n2c = Not200Club(timeout = 60)
    n2c.main()
    
    print(f'\nTotal Time to complete: {timedelta(seconds=time()-start)}s')